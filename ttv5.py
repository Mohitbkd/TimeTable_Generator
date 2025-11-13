import pandas as pd
import sys
sys.setrecursionlimit(20000)
from dataclasses import dataclass
from datetime import datetime, time
from typing import List, Dict, Tuple, Optional
import itertools, random, math
from openpyxl import load_workbook

def parse_time(val) -> time:
    """
    Parse time values in both 12-hour (e.g., '01:30 PM') and 24-hour (e.g., '13:30')
    formats into a datetime.time object.
    Returns None if input is empty or NaN.
    """
    if pd.isna(val) or str(val).strip() == "":
        return None

    # if already a time or datetime
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()

    s = str(val).strip().upper().replace('.', ':')  # normalize '1.30 PM' -> '1:30 PM'

    # ✅ Try common 12-hour and 24-hour formats
    time_formats = [
        "%I:%M %p",    # 12-hour format with minutes (e.g., "01:30 PM")
        "%I %p",       # 12-hour format without minutes (e.g., "1 PM")
        "%H:%M",       # 24-hour format (e.g., "13:30")
        "%H:%M:%S",    # 24-hour with seconds (e.g., "13:30:00")
    ]

    for fmt in time_formats:
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue

    # ✅ Fallback to pandas parser (handles edge cases)
    try:
        return pd.to_datetime(s).time()
    except Exception:
        raise ValueError(f"Invalid time format: {val!r}. Expected HH:MM / HH:MM:SS / HH:MM AM/PM")



def time_to_minutes(t: time) -> int:
    return t.hour * 60 + t.minute

def overlap(a_start: time, a_end: time, b_start: time, b_end: time) -> bool:
    return max(a_start, b_start) < min(a_end, b_end)

@dataclass(frozen=True)
class Timeslot:
    id: str
    start: time
    end: time
    @property
    def duration_min(self) -> int:
        return time_to_minutes(self.end) - time_to_minutes(self.start)

@dataclass
class Requirement:
    course_code: str
    curriculum: str
    semester: str
    section_id: str
    teacher: str    
    slots_required: int
    min_total_hours: float
    available_rooms: List[str]  # comma-split in input; empty means no room assignment
    
class TimetableCSPv2:
    def __init__(self, timeslots: List[Timeslot], requirements: List[Requirement], days: List[str], teacher_availability=None, allow_partial=False, debug=False):
        self.timeslots = timeslots
        self.requirements = requirements
        self.days = days
        self.allow_partial = allow_partial
        self.debug = debug  # Enable detailed debugging output
        self.skipped_requirements: List[Tuple[Requirement, str]] = []  # (requirement, reason)
        self.teacher_availability: Dict[str, List[Tuple[str, time, time]]] = {}
        self.allowed_teacher_pairs: Dict[str, set] = {}
        self.ta_warnings: List[str] = []

        raw_teacher_availability = teacher_availability or {}
        for teacher, periods in raw_teacher_availability.items():
            normalized: List[Tuple[str, time, time]] = []
            for day, start, end in periods:
                day_norm = str(day).strip()
                if day_norm not in self.days:
                    self.ta_warnings.append(
                        f"teacher={teacher}: day '{day_norm}' not found in DAYS sheet; skipping this availability window"
                    )
                    continue
                if start is None or end is None:
                    self.ta_warnings.append(
                        f"teacher={teacher}: availability window on {day_norm} has missing start/end; skipping"
                    )
                    continue
                if end <= start:
                    self.ta_warnings.append(
                        f"teacher={teacher}: availability window on {day_norm} has non-positive duration; skipping"
                    )
                    continue
                normalized.append((day_norm, start, end))

            if not normalized:
                continue

            allowed = set()
            for day_norm, start, end in normalized:
                for ts in self.timeslots:
                    if ts.start >= start and ts.end <= end:
                        allowed.add((day_norm, ts.id))

            if allowed:
                self.teacher_availability[teacher] = normalized
                self.allowed_teacher_pairs[teacher] = allowed
            else:
                self.ta_warnings.append(
                    f"teacher={teacher}: availability windows do not match any defined timeslot; ignoring availability"
                )

        if self.ta_warnings:
            print("[DIAG] Teacher availability warnings:")
            for msg in self.ta_warnings[:50]:
                print(f"       - {msg}")
            if len(self.ta_warnings) > 50:
                print(f"       ... and {len(self.ta_warnings) - 50} more")

        self.variables: List[Tuple[str,str,int]] = []
        self.req_index: Dict[Tuple[str,str], Requirement] = {}
        for req in requirements:
            self.req_index[(req.course_code, req.section_id)] = req
            for i in range(req.slots_required):
                self.variables.append((req.course_code, req.section_id, i))

        # Domains: each var can be placed on (day, slot_id, room_choice)
        self.domains: Dict[Tuple[str,str,int], List[Tuple[str,str,str]]] = {}
        for var in self.variables:
            c, s, _ = var
            req = self.req_index[(c, s)]
            rooms = req.available_rooms if req.available_rooms else [""]
            day_slot_pairs = list(itertools.product(self.days, [ts.id for ts in self.timeslots]))
            # Filter by teacher availability if specified for this teacher
            if req.teacher in self.allowed_teacher_pairs:
                allowed = self.allowed_teacher_pairs[req.teacher]
                day_slot_pairs = [p for p in day_slot_pairs if p in allowed]
            combos = list(itertools.product(day_slot_pairs, rooms))
            self.domains[var] = [(d, sid, r) for ((d, sid), r) in combos]

        self.ts_by_id: Dict[str, Timeslot] = {ts.id: ts for ts in self.timeslots}
        # Precompute max slot duration once
        self.max_slot_minutes = max((t.duration_min for t in self.timeslots), default=0)
        # Overlap by time (day handled in occupancy maps)
        self.conflicts: Dict[str, set] = {ts.id: {ts.id} for ts in self.timeslots}
        for a, b in itertools.combinations(self.timeslots, 2):
            if overlap(a.start, a.end, b.start, b.end):
                self.conflicts[a.id].add(b.id)
                self.conflicts[b.id].add(a.id)

        # Early domain emptiness check (fast fail)
        for var, dom in self.domains.items():
            if not dom:
                c, s, _ = var
                req = self.req_index[(c, s)]
                raise RuntimeError(
                    f"Empty domain for {var}. Likely no timeslot fits teacher availability for teacher={req.teacher}."
                )

        self.assignment: Dict[Tuple[str,str,int], Tuple[str,str,str]] = {}
        self.partial_minutes: Dict[Tuple[str,str], int] = {(c,s):0 for (c,s) in self.req_index.keys()}

        # Day-wise occupancy to avoid clashes
        self.section_busy = {d: {} for d in self.days}  # day -> section -> set(slot_ids)
        self.teacher_busy = {d: {} for d in self.days}  # day -> teacher -> set(slot_ids)
        self.room_busy    = {d: {} for d in self.days}  # day -> room -> set(slot_ids)

    def _is_virtual_room(self, room: str) -> bool:
        """Check if a room is virtual/online (no physical capacity limit)"""
        if not room:
            return False
        room_lower = room.lower().strip()
        # List of keywords that indicate virtual/online rooms
        virtual_keywords = ['online', 'virtual', 'zoom', 'teams', 'meet', 'webex', 'remote']
        return any(keyword in room_lower for keyword in virtual_keywords)
    
    def _is_feasible(self, var, val) -> bool:
        c, s, _ = var
        day, slot_id, room = val
        req = self.req_index[(c, s)]
        ts = self.ts_by_id[slot_id]
        curriculum = req.curriculum

        available_periods = self.teacher_availability.get(req.teacher, [])
        if available_periods:
            fits = any(
                day == ap[0] and ts.start >= ap[1] and ts.end <= ap[2]
                for ap in available_periods
            )
            if not fits:
                return False

        # Section clash (curriculum-aware)
        for used_slot in self.section_busy[day].get((curriculum, s), set()):
            if slot_id in self.conflicts[used_slot]:
                return False
        # Teacher clash (GLOBAL: a teacher cannot teach two places at once)
        for used_slot in self.teacher_busy[day].get(req.teacher, set()):
            if slot_id in self.conflicts[used_slot]:
                return False
        # Room clash (GLOBAL: a room cannot host two classes at once)
        # EXCEPTION: Virtual/online rooms can host unlimited classes simultaneously
        if room and not self._is_virtual_room(room):
            for used_slot in self.room_busy[day].get(room, set()):
                if slot_id in self.conflicts[used_slot]:
                    return False

        # Duration feasibility
        assigned_count = sum(1 for v in self.assignment if v[0]==c and v[1]==s)
        remaining = self.req_index[(c,s)].slots_required - assigned_count - 1
        max_slot = self.max_slot_minutes
        min_needed = int(self.req_index[(c,s)].min_total_hours * 60)
        future_possible = self.partial_minutes[(c,s)] + ts.duration_min + remaining * max_slot
        return future_possible >= min_needed

    def _place(self, var, val):
        # confirms the assignment is okay, during backtracking search
        c, s, _ = var
        day, slot_id, room = val
        req = self.req_index[(c, s)]
        curriculum = req.curriculum

        self.assignment[var] = val
        self.partial_minutes[(c,s)] += self.ts_by_id[slot_id].duration_min

        self.section_busy[day].setdefault((curriculum, s), set()).add(slot_id)
        self.teacher_busy[day].setdefault(req.teacher, set()).add(slot_id)
        # Only track physical room occupancy (virtual rooms have unlimited capacity)
        if room and not self._is_virtual_room(room):
            self.room_busy[day].setdefault(room, set()).add(slot_id)
    
    def _remove(self, var, val):
        # undoes the assignment, during backtracking search
        c, s, _ = var
        day, slot_id, room = val
        req = self.req_index[(c, s)]
        curriculum = req.curriculum

        del self.assignment[var]
        self.partial_minutes[(c,s)] -= self.ts_by_id[slot_id].duration_min

        self.section_busy[day][(curriculum, s)].remove(slot_id)
        if not self.section_busy[day][(curriculum, s)]:
            del self.section_busy[day][(curriculum, s)]

        self.teacher_busy[day][req.teacher].remove(slot_id)
        if not self.teacher_busy[day][req.teacher]:
            del self.teacher_busy[day][req.teacher]

        # Only remove physical room occupancy (virtual rooms aren't tracked)
        if room and not self._is_virtual_room(room):
            self.room_busy[day][room].remove(slot_id)
            if not self.room_busy[day][room]:
                del self.room_busy[day][room]

    def _mrv(self): # Implements the Minimum Remaining Values
        best, best_size = None, math.inf
        # Get skipped vars if in partial mode
        skipped = getattr(self, 'skipped_vars', set())
        
        for var in self.variables:
            if var in self.assignment or var in skipped:
                continue
            feasible_count = 0
            for val in self.domains[var]:
                if self._is_feasible(var, val):
                    feasible_count += 1
            if feasible_count < best_size:
                best, best_size = var, feasible_count
                if best_size == 0:
                    break
        return best

    def _order_values(self, var, candidates):
        # Heuristic:
        # 1) Prefer days NOT already used by this course-section (spread across days)
        # 2) Prefer slots with fewer conflicts
        # 3) Prefer earlier in the week/day
        c, s, slot_idx = var
        req = self.req_index[(c, s)]

        day_rank = {d:i for i,d in enumerate(self.days)}
        
        # Get days already used by this course-section
        used_days = set()
        for (cc, ss, _), val in self.assignment.items():
            if cc == c and ss == s and val is not None:
                used_days.add(val[0])  # val[0] is day

        def total_load(day, slot_id):
            # count how many occupancies currently use this slot_id on this day
            sec_load = sum(1 for used in self.section_busy[day].values() if slot_id in used)
            tch_load = sum(1 for used in self.teacher_busy[day].values() if slot_id in used)
            rm_load = sum(1 for used in self.room_busy[day].values() if slot_id in used)
            return sec_load + tch_load + rm_load

        def static_conflicts(day, slot_id, room):
            # rough estimate: overlaps with already used slots for this section/teacher/room on the same day
            sec_conf = sum(1 for used in self.section_busy[day].get((req.curriculum, s), set()) if slot_id in self.conflicts[used])
            teach_conf = sum(1 for used in self.teacher_busy[day].get(req.teacher, set()) if slot_id in self.conflicts[used])
            room_conf = 0
            if room:
                room_conf = sum(1 for used in self.room_busy[day].get(room, set()) if slot_id in self.conflicts[used])
            return sec_conf + teach_conf + room_conf

        def score(val):
            day, slot_id, room = val
            ts = self.ts_by_id[slot_id]
            # Prefer days NOT already used (0 if new day, 1 if already used)
            day_already_used = 1 if day in used_days else 0
            return (
                day_already_used,                         # NEW: prefer different days (0 < 1)
                static_conflicts(day, slot_id, room),     # fewer conflicts first
                total_load(day, slot_id),                 # lower load better
                day_rank.get(day, 999),                   # earlier day
                ts.start                                  # earlier time
            )

        return sorted(candidates, key=score)

    def _backtrack(self, max_attempts_per_var=1000) -> bool:
        # Count only non-skipped variables
        if not hasattr(self, 'skipped_vars'):
            self.skipped_vars = set()
        
        non_skipped_vars = [v for v in self.variables if v not in self.skipped_vars]
        assigned_count = len([v for v in non_skipped_vars if v in self.assignment])
        
        if assigned_count == len(non_skipped_vars):
            # final check: min hours for non-skipped requirements
            for (c, s), req in self.req_index.items():
                # Skip if this requirement was skipped
                if any(v for v in self.skipped_vars if v[0] == c and v[1] == s):
                    continue
                if self.partial_minutes[(c, s)] < int(req.min_total_hours*60):
                    return False
            return True

        var = self._mrv()
        # Progress tracking
        if assigned_count % 10 == 0 and assigned_count != getattr(self, '_last_progress', -1):
            print(f"[PROGRESS] Assigned {assigned_count}/{len(non_skipped_vars)} variables")
            self._last_progress = assigned_count

        if var is None:
            # In partial mode, if no variable found but we have skipped some, that's ok
            if self.allow_partial and self.skipped_vars:
                return True
            return False
            
        # Track attempts per variable
        if not hasattr(self, 'attempt_count'):
            self.attempt_count = {}
        self.attempt_count[var] = self.attempt_count.get(var, 0) + 1
        
        # Debug: Show which requirement is being processed
        if self.debug and self.attempt_count[var] == 1:
            c, s, i = var
            req = self.req_index[(c, s)]
            print(f"[DEBUG] Processing: course={req.course_code} section={req.section_id} teacher={req.teacher} slot_index={i}/{req.slots_required}")
        
        # Debug: Warn if taking many attempts
        if self.debug and self.attempt_count[var] in [100, 250, 400]:
            c, s, i = var
            req = self.req_index[(c, s)]
            print(f"[DEBUG] Struggling with: course={req.course_code} section={req.section_id} teacher={req.teacher} (attempt {self.attempt_count[var]})")
        
        # If stuck on a variable for too long in partial mode, skip it
        if self.allow_partial and self.attempt_count[var] > max_attempts_per_var:
            c, s, i = var
            req = self.req_index[(c, s)]
            
            # Determine reason for failure
            feasible = [v for v in self.domains[var] if self._is_feasible(var, v)]
            if len(feasible) == 0:
                reason = "No feasible timeslots available (teacher busy, section conflict, or room occupied)"
            else:
                reason = f"Could not place after {max_attempts_per_var} attempts (likely over-constrained or conflicting requirements)"
            
            # Skip all slots for this requirement
            req_key = (c, s)
            if req_key not in [r[0] for r in self.skipped_requirements]:
                self.skipped_requirements.append((req, reason))
                print(f"[SKIP] course={c} section={s} teacher={req.teacher}: {reason}")
            
            # Mark all variables for this requirement as skipped
            vars_to_skip = [(cc, ss, ii) for (cc, ss, ii) in self.variables if cc == c and ss == s]
            for v in vars_to_skip:
                self.skipped_vars.add(v)
            
            return self._backtrack(max_attempts_per_var)  # Continue with next variable

        if var is None:
            return False

        candidates = [v for v in self.domains[var] if self._is_feasible(var, v)]
        # In partial mode with large datasets, if no candidates, skip immediately
        if self.allow_partial and len(candidates) == 0:
            c, s, i = var
            req = self.req_index[(c, s)]
            req_key = (c, s)
            if req_key not in [r[0] for r in self.skipped_requirements]:
                self.skipped_requirements.append((req, "No feasible timeslots available (teacher busy, section conflict, or room occupied)"))
                print(f"[SKIP] course={req.course_code} section={req.section_id} teacher={req.teacher}: No feasible slots")
            vars_to_skip = [(cc, ss, ii) for (cc, ss, ii) in self.variables if cc == c and ss == s]
            for v in vars_to_skip:
                self.skipped_vars.add(v)
            return self._backtrack(max_attempts_per_var)
        
        # Order values by LCV to reduce backtracking
        candidates = self._order_values(var, candidates)
        
        # For large datasets, limit candidate exploration to top 20 options
        if self.allow_partial and len(candidates) > 20:
            if self.debug:
                c, s, i = var
                req = self.req_index[(c, s)]
                print(f"[DEBUG] Limiting candidates from {len(candidates)} to 20 for: course={req.course_code} section={req.section_id}")
            candidates = candidates[:20]

        for val in candidates:
            self._place(var, val)
            if self._backtrack():
                return True
            self._remove(var, val)

        return False

    def solve(self, seed: int = 123):
        random.seed(seed)
        if not self._backtrack():
            if self.allow_partial:
                print("\n[PARTIAL SOLUTION] Some requirements could not be scheduled.")
            else:
                print("Debug Info: Variables with 0 feasible options:")
                zero_count = 0
                for var in self.variables:
                    feasible = [v for v in self.domains[var] if self._is_feasible(var, v)]
                    if len(feasible) == 0:
                        print(f"  {var}: 0 feasible options")
                        zero_count += 1
                if zero_count == 0:
                    print("  None found; check for section/teacher/room conflicts or hour requirements.")
                raise RuntimeError("No feasible timetable found with current inputs/constraints.")
        
        # Filter out skipped variables
        if not hasattr(self, 'skipped_vars'):
            self.skipped_vars = set()
        valid_assignments = {k: v for k, v in self.assignment.items() if k not in self.skipped_vars}
        
        # Report skipped requirements
        if self.skipped_requirements:
            print(f"\n{'='*80}")
            print(f"[UNSCHEDULED REQUIREMENTS] {len(self.skipped_requirements)} requirement(s) could not be scheduled:")
            print(f"{'='*80}")
            for req, reason in self.skipped_requirements:
                print(f"\n  Course: {req.course_code}")
                print(f"  Section: {req.section_id}")
                print(f"  Teacher: {req.teacher}")
                print(f"  Curriculum: {req.curriculum}")
                print(f"  Semester: {req.semester}")
                print(f"  Slots Required: {req.slots_required}")
                print(f"  Min Hours: {req.min_total_hours}")
                print(f"  Reason: {reason}")
                print(f"  {'-'*78}")
            print(f"\n[RECOMMENDATIONS]")
            print("  1. Reduce 'slots_required' for these courses in Excel")
            print("  2. Expand teacher availability in TEACHER_AVAILABILITY sheet")
            print("  3. Assign different teachers who have more availability")
            print("  4. Check for duplicate or conflicting requirements")
            print(f"{'='*80}\n")
        
        return valid_assignments
    
def read_input_v2(xlsx_path: str):
    xls = pd.ExcelFile(xlsx_path)
    window = pd.read_excel(xlsx_path, "WINDOW")
    slots  = pd.read_excel(xlsx_path, "TIMESLOTS")
    reqdf  = pd.read_excel(xlsx_path, "REQUIREMENTS")
    daysdf = pd.read_excel(xlsx_path, "DAYS") if "DAYS" in xls.sheet_names else pd.DataFrame({"day":["Mon","Tue","Wed","Thu","Fri"]})

    start_date = str(window.iloc[0]["start_date"])
    end_date   = str(window.iloc[0]["end_date"])

    # ✅ Build timeslots (skip invalid)
    timeslots = []
    for r in slots.itertuples(index=False):
        sid = str(r.slot_id)
        start_time_str = str(r.start_time).strip()
        end_time_str = str(r.end_time).strip()

        # If start_time contains space and end_time is empty, split it
        if ' ' in start_time_str and not end_time_str:
            parts = start_time_str.split()
            if len(parts) >= 4:
                start_str = parts[0] + ' ' + parts[1]  # e.g., '01:30 PM'
                end_str = parts[2] + ' ' + parts[3]    # e.g., '03:00 PM'
            else:
                start_str, end_str = start_time_str, ''
        else:
            start_str, end_str = start_time_str, end_time_str

        st = parse_time(start_str)
        et = parse_time(end_str)
        if st is None or et is None:
            print(f"[DIAG] Skipping TIMESLOTS row with empty time: slot_id={sid} start={r.start_time} end={r.end_time}")
            continue
        if et <= st:
            print(f"[DIAG] Skipping TIMESLOTS row with non-positive duration: slot_id={sid} start={st} end={et}")
            continue
        timeslots.append(Timeslot(sid, st, et))

    # ✅ Build requirements
    reqs = []
    for r in reqdf.itertuples(index=False):
        rooms = [x.strip() for x in str(getattr(r, "available_rooms", "") or "").split(",") if x.strip()]
        cc = str(r.course_code).strip()
        cur = str(r.curriculum).strip()
        sem = str(r.semester).strip()
        sec = str(r.section_id).strip()
        tch = str(r.teacher).strip()
        reqs.append(Requirement(
            course_code=cc,
            curriculum=cur,
            semester=sem,
            section_id=sec,
            teacher=tch,
            slots_required=int(r.slots_required),
            min_total_hours=float(r.min_total_hours),
            available_rooms=rooms
        ))

    # ✅ Build days
    days = [str(d).strip() for d in daysdf["day"].tolist()]

    # ✅ Breaks
    breaks = []
    if "BREAKS" in xls.sheet_names:
        brdf = pd.read_excel(xlsx_path, "BREAKS")
        for r in brdf.itertuples(index=False):
            bf = parse_time(getattr(r, "break_from", "") or "")
            bt = parse_time(getattr(r, "break_to", "") or "")
            if bf is None or bt is None:
                print(f"⚠️ Skipping BREAKS row with invalid time: {r}")
                continue
            breaks.append({
                "curriculum": str(getattr(r, "curriculum", "")).strip(),
                "semester": str(getattr(r, "semester", "")).strip(),
                "section": str(getattr(r, "section_id", "")).strip(),
                "day": str(getattr(r, "day", "")).strip(),
                "break_from": bf,
                "break_to": bt
            })

    # ✅ Teacher availability (support multiple days like "Mon / Wed")
    teacher_availability = {}
    if "TEACHER_AVAILABILITY" in xls.sheet_names:
        tavdf = pd.read_excel(xlsx_path, "TEACHER_AVAILABILITY")
        for r in tavdf.itertuples(index=False):
            teacher = str(r.teacher).strip()
            raw_days = str(r.day).strip()

            # Split multiple days like "Mon / Wed / Fri"
            day_list = [d.strip() for d in raw_days.replace("\\", "/").split("/") if d.strip()]

            available_from_str = str(r.available_from).strip()
            available_to_str = str(r.available_to).strip()

            # If available_from contains space and available_to is empty, split it
            if ' ' in available_from_str and not available_to_str:
                parts = available_from_str.split()
                if len(parts) >= 2:
                    start_str = parts[0] + ' ' + parts[1]  # e.g., '01:30 PM'
                    end_str = parts[2] + ' ' + parts[3] if len(parts) >= 4 else parts[2]  # e.g., '03:00 PM'
                else:
                    start_str, end_str = available_from_str, ''
            else:
                start_str, end_str = available_from_str, available_to_str

            start = parse_time(start_str)
            end = parse_time(end_str)
            if start is None or end is None:
                print(f"⚠️ Skipping TEACHER_AVAILABILITY row with invalid time: {r}")
                continue

            # Add availability for each day
            for day in day_list:
                teacher_availability.setdefault(teacher, []).append((day, start, end))

    return {
        "start_date": start_date,
        "end_date": end_date,
        "timeslots": timeslots,
        "requirements": reqs,
        "days": days,
        "breaks": breaks,
        "teacher_availability": teacher_availability
    }

def is_teacher_available(teacher_availability, teacher: str, day: str, slot_start, slot_end) -> bool:
    """
    Check if a teacher is available on a given day and timeslot.
    """
    # If no availability specified for this teacher → assume always available
    availability = [a for a in teacher_availability if a["teacher"] == teacher and a["day"] == day]
    if not availability:
        return True  

    for a in availability:
        if slot_start >= a["available_from"] and slot_end <= a["available_to"]:
            return True
    return False

def export_to_template(assignments, engine, start_date, end_date, output_xlsx, template_xlsx, break_start="12:00", break_end="12:30", holidays=None, skipped_requirements=None):
    if holidays is None:
        holidays = []  # list of days like ["Sunday"]
    if skipped_requirements is None:
        skipped_requirements = []

    # Build rows grouped by course-section (up to 5 meetings per row)
    day_rank = {d: i for i, d in enumerate(engine.days)}
    from collections import defaultdict
    buckets = defaultdict(list)

    # Normal assignments
    assignment_count = 0
    for (course, section, _), (day, slot_id, room) in assignments.items():
        ts = engine.ts_by_id[slot_id]
        req = engine.req_index[(course, section)]
        buckets[(course, section)].append({
            "course": course,
            "section": section,
            "teacher": req.teacher,
            "curriculum": req.curriculum,
            "semester": req.semester,
            "day": day,
            "time_from": ts.start.strftime("%H:%M:%S"),
            "time_to": ts.end.strftime("%H:%M:%S"),
            "room": room or "",
            "start": ts.start.strftime("%H:%M")
        })
        assignment_count += 1
    
    print(f"[EXPORT] Processing {assignment_count} scheduled assignments")

    # Inject break slots (for all courses/sections, common for each day)
    sections = {(req.curriculum, req.semester, req.section_id) for req in engine.requirements}

    for (curr, sem, sec) in sections:
        for day in engine.days:
            if day in holidays:
                continue
            # bucket key without course
            buckets[(None, sec)].append({
                "course": "",   # no course
                "section": sec,
                "teacher": "",  # no teacher
                "curriculum": curr,
                "semester": sem,
                "day": day,
                "time_from": f"{break_start}:00",
                "time_to": f"{break_end}:00",
                "room": "BREAK",
                "start": break_start
            })


    # Build final rows
    rows = []
    for key, items in buckets.items():
        items.sort(key=lambda x: (day_rank.get(x["day"], 999), x["start"]))
        for off in range(0, len(items), 5):
            chunk = items[off:off + 5]
            base = chunk[0]
            row = {
                "STARTDATE": start_date, "ENDDATE": end_date,
                "CURRICULUM": base["curriculum"], "COURSE": base["course"],
                "SEMESTER": base["semester"], "SECTION": base["section"], "TEACHER": base["teacher"],
            }
            for i in range(5):
                dcol, tfcol, ttcol, rcol, lcol = (
                    f"DAY{i+1}", f"DAY{i+1}_TIME_FROM", f"DAY{i+1}_TIME_TO", f"DAY{i+1}_ROOM", f"DAY{i+1}_LINK"
                )
                if i < len(chunk):
                    e = chunk[i]
                    row[dcol], row[tfcol], row[ttcol], row[rcol], row[lcol] = (
                        e["day"], e["time_from"], e["time_to"], e["room"], ""
                    )
                else:
                    row[dcol] = row[tfcol] = row[ttcol] = row[rcol] = row[lcol] = ""
            rows.append(row)

    out_df = pd.DataFrame(rows, columns=[
        "STARTDATE", "ENDDATE", "CURRICULUM", "COURSE", "SEMESTER", "SECTION", "TEACHER",
        "DAY1", "DAY1_TIME_FROM", "DAY1_TIME_TO", "DAY1_ROOM", "DAY1_LINK",
        "DAY2", "DAY2_TIME_FROM", "DAY2_TIME_TO", "DAY2_ROOM", "DAY2_LINK",
        "DAY3", "DAY3_TIME_FROM", "DAY3_TIME_TO", "DAY3_ROOM", "DAY3_LINK",
        "DAY4", "DAY4_TIME_FROM", "DAY4_TIME_TO", "DAY4_ROOM", "DAY4_LINK",
        "DAY5", "DAY5_TIME_FROM", "DAY5_TIME_TO", "DAY5_ROOM", "DAY5_LINK"
    ])

    wb = load_workbook(template_xlsx)
    ws = wb["TimeTable"]
    headers = [c.value for c in ws[1]]

    # Remove all sheets except "TimeTable"
    for sheet_name in list(wb.sheetnames):
        if sheet_name != "TimeTable":
            del wb[sheet_name]

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for _, r in out_df.iterrows():
        ws.append([r.get(h, "") for h in headers])

    # Add Unscheduled sheet if there are skipped requirements
    if skipped_requirements:
        unscheduled_rows = []
        for req, reason in skipped_requirements:
            unscheduled_rows.append({
                "Course": req.course_code,
                "Section": req.section_id,
                "Teacher": req.teacher,
                "Curriculum": req.curriculum,
                "Semester": req.semester,
                "Slots Required": req.slots_required,
                "Min Hours": req.min_total_hours,
                "Available Rooms": ", ".join(req.available_rooms) if req.available_rooms else "",
                "Reason": reason
            })
        
        unscheduled_df = pd.DataFrame(unscheduled_rows)
        
        # Create new sheet for unscheduled requirements
        ws_unscheduled = wb.create_sheet("Unscheduled")
        
        # Write headers
        for col_idx, col_name in enumerate(unscheduled_df.columns, start=1):
            ws_unscheduled.cell(row=1, column=col_idx, value=col_name)
        
        # Write data
        for row_idx, row_data in enumerate(unscheduled_df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws_unscheduled.cell(row=row_idx, column=col_idx, value=value)
        
        print(f"[INFO] Added 'Unscheduled' sheet with {len(unscheduled_rows)} requirement(s)")

    wb.save(output_xlsx)
    return output_xlsx

# ------------------- Main Execution -------------------

if __name__ == "__main__":
    """Main execution block - only runs when script is executed directly"""
    import os
    
    REQUIRED_SHEETS = ["WINDOW", "TIMESLOTS", "REQUIREMENTS", "DAYS"]
    
    input_xlsx = "InputData_v2.xlsx"
    
    print("Using:", input_xlsx)

    # Read breaks sheet
    breaks_df = pd.read_excel(input_xlsx, sheet_name="BREAKS")

    # Function to fetch break time
    def get_break_time(curriculum, semester, section_id, day):
        match = breaks_df[
            (breaks_df["curriculum"] == curriculum) &
            (breaks_df["semester"] == semester) &
            (breaks_df["section_id"] == section_id) &
            (breaks_df["day"] == day)
        ]
        if not match.empty:
            return match.iloc[0]["break_from"], match.iloc[0]["break_to"]
        return None, None

    # Quick validation
    # ✅ Quick validation
    if not os.path.exists(input_xlsx):
        raise FileNotFoundError(f"❌ File not found: {input_xlsx}")
        
    xls = pd.ExcelFile(input_xlsx)
    missing = [s for s in REQUIRED_SHEETS if s not in xls.sheet_names]

    if missing:
        raise ValueError(
            f"❌ Input file is missing required sheets: {missing}. "
            f"Found sheets: {xls.sheet_names}"
        )
    else:
        print("✅ All required sheets are present!")


    # (Optional) quick peek so you can confirm you uploaded the right file
    # display(pd.read_excel(input_xlsx, "WINDOW").head())
    # display(pd.read_excel(input_xlsx, "TIMESLOTS").head())
    # display(pd.read_excel(input_xlsx, "REQUIREMENTS").head())
    # display(pd.read_excel(input_xlsx, "DAYS").head())


    # Now run the pipeline with this uploaded input
    data = read_input_v2(input_xlsx)

    engine = TimetableCSPv2(
        data["timeslots"], 
        data["requirements"], 
        data["days"],
        data['teacher_availability'],  # <-- pass teacher availability here
        allow_partial=True,  # <-- Enable partial solution mode (skip impossible requirements)
        debug=True  # <-- Enable debug mode to track which requirements are being processed
    )

    print("Generating timetable... This may take a few seconds.")


    # ---------- Pre-solve diagnostics ----------
    try:
        print(f"[DIAG] Loaded days={len(engine.days)} timeslots={len(engine.timeslots)} requirements={len(engine.requirements)}")
        days_set = set(engine.days)
        max_slot_min = max(ts.duration_min for ts in engine.timeslots) if engine.timeslots else 0

        # Check teacher availability day labels
        ta_day_issues = []
        for t, periods in (engine.teacher_availability or {}).items():
            for (d, _s, _e) in periods:
                if d not in days_set:
                    ta_day_issues.append((t, d))
        if ta_day_issues:
            print("[DIAG] Teacher availability contains days not in DAYS sheet:")
            for t, d in ta_day_issues[:20]:
                print(f"       - teacher={t} day={d} (not in {sorted(days_set)})")
            if len(ta_day_issues) > 20:
                print(f"       ... and {len(ta_day_issues)-20} more")

        # For each requirement: check overlap options and hours feasibility
        print("[DIAG] Requirement feasibility quick-check:")
        for req in engine.requirements[:2000]:
            # availability overlap count
            options = 0
            periods = engine.teacher_availability.get(req.teacher, [])
            for d in engine.days:
                for ts in engine.timeslots:
                    if periods:
                        fits = any(d == ap[0] and ts.start >= ap[1] and ts.end <= ap[2] for ap in periods)
                        if not fits:
                            continue
                    options += 1

            min_needed = int(req.min_total_hours * 60)
            max_possible = req.slots_required * max_slot_min
            msgs = []
            if options == 0:
                msgs.append("no (day,timeslot) fits teacher availability")
            if max_possible < min_needed:
                msgs.append(f"insufficient total minutes: need {min_needed}, have at most {max_possible} from {req.slots_required} slot(s)")
            if msgs:
                print(
                    f"       - course={req.course_code} section={req.section_id} teacher={req.teacher} "
                    f"curriculum={req.curriculum} semester={req.semester}: " + "; ".join(msgs)
                )

        # Teacher supply vs demand diagnostics
        try:
            teacher_slot_demand: Dict[str, int] = {}
            teacher_minute_demand: Dict[str, int] = {}
            for req in engine.requirements:
                teacher_slot_demand[req.teacher] = teacher_slot_demand.get(req.teacher, 0) + req.slots_required
                teacher_minute_demand[req.teacher] = teacher_minute_demand.get(req.teacher, 0) + int(req.min_total_hours * 60)

            if teacher_slot_demand:
                slots_per_week = len(engine.days) * len(engine.timeslots)
                minutes_per_day = sum(ts.duration_min for ts in engine.timeslots)
                full_supply_minutes = minutes_per_day * len(engine.days)
                ts_minutes = {ts.id: ts.duration_min for ts in engine.timeslots}

                over_subscribed = []
                for teacher, slot_demand in teacher_slot_demand.items():
                    minute_demand = teacher_minute_demand.get(teacher, 0)
                    if teacher in engine.allowed_teacher_pairs:
                        allowed_pairs = engine.allowed_teacher_pairs[teacher]
                        slot_supply = len(allowed_pairs)
                        minute_supply = sum(ts_minutes[sid] for (_day, sid) in allowed_pairs)
                    else:
                        slot_supply = slots_per_week
                        minute_supply = full_supply_minutes

                    if slot_demand > slot_supply or minute_demand > minute_supply:
                        over_subscribed.append((teacher, slot_demand, slot_supply, minute_demand, minute_supply))

                if over_subscribed:
                    print("[DIAG] Teacher supply vs demand issues:")
                    for teacher, sd, ss, md, ms in over_subscribed[:20]:
                        print(
                            "       - teacher={teacher} demand_slots={sd} supply_slots={ss} demand_minutes={md} supply_minutes={ms}".format(
                                teacher=teacher,
                                sd=sd,
                                ss=ss,
                                md=md,
                                ms=ms,
                            )
                        )
                    if len(over_subscribed) > 20:
                        print(f"       ... and {len(over_subscribed) - 20} more")
        except Exception as teacher_diag_err:
            print(f"[DIAG] Skipped teacher supply diagnostics due to error: {teacher_diag_err}")
    except Exception as diag_err:
        print(f"[DIAG] Skipped diagnostics due to error: {diag_err}")


    # ---------- Section capacity diagnostics ----------
    try:
        weekly_supply = len(engine.days) * len(engine.timeslots)
        demand_by_section = {}
        for req in engine.requirements:
            demand_by_section[req.section_id] = demand_by_section.get(req.section_id, 0) + req.slots_required
        over = [(sec, dem) for sec, dem in demand_by_section.items() if dem > weekly_supply]
        if over:
            print("[DIAG] Section weekly slot demand exceeds supply:")
            for sec, dem in over:
                print(f"       - section={sec} demand_slots={dem} > weekly_supply={weekly_supply} (len(DAYS)={len(engine.days)} * slots_per_day={len(engine.timeslots)})")
    except Exception as diag_err2:
        print(f"[DIAG] Skipped section capacity diagnostics due to error: {diag_err2}")


    assignments = engine.solve(seed=123)

    # Excel template path
    template_xlsx = "TimeTableImport_SIS.xlsx"


    if not os.path.exists(template_xlsx):
        raise FileNotFoundError(f"❌ Template file not found: {template_xlsx}")

    # 🔹 Output file path
    out_path = "GeneratedTimetable.xlsx"

    # Export timetable
    export_to_template(
        assignments,
        engine,
        data["start_date"],
        data["end_date"],
        out_path,
        template_xlsx,
        skipped_requirements=engine.skipped_requirements  # Pass unscheduled requirements
    )

    print(f"✅ Done! Timetable exported successfully to: {out_path}")
