import pandas as pd
from dataclasses import dataclass
from datetime import datetime, time
from typing import List, Dict, Tuple, Optional
import itertools, random, math
from openpyxl import load_workbook

def parse_time(val) -> time:
    if pd.isna(val) or val == "":
        return None
    # if already a time or datetime
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    s = str(val).strip()
    # try common formats
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p"):
        try:
            return datetime.strptime(s, fmt).time()
        except Exception:
            pass
    # fallback to pandas parsing
    try:
        return pd.to_datetime(s).time()
    except Exception:
        raise ValueError(f"Invalid time format: {val!r}. Expected HH:MM or HH:MM:SS")


def time_to_minutes(t: time) -> int:
    return t.hour * 60 + t.minute

def overlap(a_start: time, a_end: time, b_start: time, b_end: time) -> bool:
    return max(a_start, b_start) < min(a_end, b_end)

@dataclass(frozen=True)
class Timeslot:
    id: str
    start: time
    end: time
    @property
    def duration_min(self) -> int:
        return time_to_minutes(self.end) - time_to_minutes(self.start)

@dataclass
class Requirement:
    course_code: str
    curriculum: str
    semester: str
    section_id: str
    teacher: str    
    slots_required: int
    min_total_hours: float
    available_rooms: List[str]  # comma-split in input; empty means no room assignment
    
class TimetableCSPv2:
    def __init__(self, timeslots: List[Timeslot], requirements: List[Requirement], days: List[str], teacher_availability=None):
        self.timeslots = timeslots
        self.requirements = requirements
        self.days = days
        self.teacher_availability = teacher_availability or {}  # <-- Add this line

        self.variables: List[Tuple[str,str,int]] = []
        self.req_index: Dict[Tuple[str,str], Requirement] = {}
        for req in requirements:
            self.req_index[(req.course_code, req.section_id)] = req
            for i in range(req.slots_required):
                self.variables.append((req.course_code, req.section_id, i))

        # Domains: each var can be placed on (day, slot_id, room_choice)
        self.domains: Dict[Tuple[str,str,int], List[Tuple[str,str,str]]] = {}
        for var in self.variables:
            c, s, _ = var
            req = self.req_index[(c, s)]
            rooms = req.available_rooms if req.available_rooms else [""]
            combos = list(itertools.product(self.days, [ts.id for ts in self.timeslots], rooms))
            self.domains[var] = [(d, sid, r) for (d, sid, r) in combos]

        self.ts_by_id: Dict[str, Timeslot] = {ts.id: ts for ts in self.timeslots}
        # Overlap by time (day handled in occupancy maps)
        self.conflicts: Dict[str, set] = {ts.id: {ts.id} for ts in self.timeslots}
        for a, b in itertools.combinations(self.timeslots, 2):
            if overlap(a.start, a.end, b.start, b.end):
                self.conflicts[a.id].add(b.id)
                self.conflicts[b.id].add(a.id)

        self.assignment: Dict[Tuple[str,str,int], Tuple[str,str,str]] = {}
        self.partial_minutes: Dict[Tuple[str,str], int] = {(c,s):0 for (c,s) in self.req_index.keys()}

        # Day-wise occupancy to avoid clashes
        self.section_busy = {d: {} for d in self.days}  # day -> section -> set(slot_ids)
        self.teacher_busy = {d: {} for d in self.days}  # day -> teacher -> set(slot_ids)
        self.room_busy    = {d: {} for d in self.days}  # day -> room -> set(slot_ids)

    def _is_feasible(self, var, val) -> bool:
        c, s, _ = var
        day, slot_id, room = val
        req = self.req_index[(c, s)]
        ts = self.ts_by_id[slot_id]
      
        available_periods = self.teacher_availability.get(req.teacher, [])
        if available_periods:
            # Only allow if the timeslot fits within any available period for this day
            fits = any(
                day == ap[0] and ts.start >= ap[1] and ts.end <= ap[2]
                for ap in available_periods
            )
            if not fits:
                return False

        # Section clash
        for used_slot in self.section_busy[day].get(s, set()):
            if slot_id in self.conflicts[used_slot]:
                return False
        # Teacher clash
        for used_slot in self.teacher_busy[day].get(req.teacher, set()):
            if slot_id in self.conflicts[used_slot]:
                return False
        # Room clash
        if room:
            for used_slot in self.room_busy[day].get(room, set()):
                if slot_id in self.conflicts[used_slot]:
                    return False

        # Duration feasibility (can we still reach min_total_hours?)
        assigned_count = sum(1 for v in self.assignment if v[0]==c and v[1]==s)
        remaining = self.req_index[(c,s)].slots_required - assigned_count - 1
        max_slot = max(t.duration_min for t in self.timeslots)
        min_needed = int(self.req_index[(c,s)].min_total_hours * 60)
        future_possible = self.partial_minutes[(c,s)] + ts.duration_min + remaining * max_slot
        return future_possible >= min_needed

    def _place(self, var, val):
        c, s, _ = var
        day, slot_id, room = val
        req = self.req_index[(c, s)]
        ts = self.ts_by_id[slot_id]

        self.assignment[var] = val
        self.partial_minutes[(c,s)] += ts.duration_min

        self.section_busy[day].setdefault(s, set()).add(slot_id)
        self.teacher_busy[day].setdefault(req.teacher, set()).add(slot_id)
        if room:
            self.room_busy[day].setdefault(room, set()).add(slot_id)

    def _remove(self, var, val):
        c, s, _ = var
        day, slot_id, room = val
        req = self.req_index[(c, s)]
        ts = self.ts_by_id[slot_id]

        del self.assignment[var]
        self.partial_minutes[(c,s)] -= ts.duration_min

        self.section_busy[day][s].remove(slot_id)
        if not self.section_busy[day][s]:
            del self.section_busy[day][s]

        self.teacher_busy[day][req.teacher].remove(slot_id)
        if not self.teacher_busy[day][req.teacher]:
            del self.teacher_busy[day][req.teacher]

        if room:
            self.room_busy[day][room].remove(slot_id)
            if not self.room_busy[day][room]:
                del self.room_busy[day][room]

    def _mrv(self):
        best, best_size = None, math.inf
        for var in self.variables:
            if var in self.assignment:
                continue
            feasible_count = 0
            for val in self.domains[var]:
                if self._is_feasible(var, val):
                    feasible_count += 1
                    if feasible_count >= best_size:
                        break
            if feasible_count < best_size:
                best, best_size = var, feasible_count
                if best_size == 0:
                    break
        return best

    def _backtrack(self) -> bool:
        if len(self.assignment) == len(self.variables):
            # final check: min hours
            for (c, s), req in self.req_index.items():
                if self.partial_minutes[(c, s)] < int(req.min_total_hours*60):
                    return False
            return True

        var = self._mrv()
        if var is None:
            return False

        candidates = [v for v in self.domains[var] if self._is_feasible(var, v)]
        random.shuffle(candidates)

        for val in candidates:
            self._place(var, val)
            if self._backtrack():
                return True
            self._remove(var, val)

        return False

    def solve(self, seed: int = 123):
        random.seed(seed)
        if not self._backtrack():
            raise RuntimeError("No feasible timetable found with current inputs/constraints.")
        return dict(self.assignment)
    
def read_input_v2(xlsx_path: str):
    xls = pd.ExcelFile(xlsx_path)
    window = pd.read_excel(xlsx_path, "WINDOW")
    slots  = pd.read_excel(xlsx_path, "TIMESLOTS")
    reqdf  = pd.read_excel(xlsx_path, "REQUIREMENTS")
    daysdf = pd.read_excel(xlsx_path, "DAYS") if "DAYS" in xls.sheet_names else pd.DataFrame({"day":["Mon","Tue","Wed","Thu","Fri"]})

    start_date = str(window.iloc[0]["start_date"])
    end_date   = str(window.iloc[0]["end_date"])

    # ✅ Build timeslots
    timeslots = [
        Timeslot(str(r.slot_id), parse_time(str(r.start_time)), parse_time(str(r.end_time)))
        for r in slots.itertuples(index=False)
    ]

    # ✅ Build requirements
    reqs = []
    for r in reqdf.itertuples(index=False):
        rooms = [x.strip() for x in str(getattr(r, "available_rooms", "") or "").split(",") if x.strip()]
        reqs.append(Requirement(
            course_code=str(r.course_code),
            curriculum=str(r.curriculum),
            semester=str(r.semester),
            section_id=str(r.section_id),
            teacher=str(r.teacher),
            slots_required=int(r.slots_required),
            min_total_hours=float(r.min_total_hours),
            available_rooms=rooms
        ))

    # ✅ Build days
    days = [str(d) for d in daysdf["day"].tolist()]

    # ✅ Breaks
    breaks = []
    if "BREAKS" in xls.sheet_names:
        brdf = pd.read_excel(xlsx_path, "BREAKS")
        for r in brdf.itertuples(index=False):
            bf = parse_time(getattr(r, "break_from", "") or "")
            bt = parse_time(getattr(r, "break_to", "") or "")
            if bf is None or bt is None:
                print(f"⚠️ Skipping BREAKS row with invalid time: {r}")
                continue
            breaks.append({
                "curriculum": str(getattr(r, "curriculum", "")).strip(),
                "semester": str(getattr(r, "semester", "")).strip(),
                "section": str(getattr(r, "section_id", "")).strip(),
                "day": str(getattr(r, "day", "")).strip(),
                "break_from": bf,
                "break_to": bt
            })

    # ✅ Teacher availability
    teacher_availability = {}
    if "TEACHER_AVAILABILITY" in xls.sheet_names:
        tavdf = pd.read_excel(xlsx_path, "TEACHER_AVAILABILITY")
        for r in tavdf.itertuples(index=False):
            teacher = str(r.teacher).strip()
            day = str(r.day).strip()
            start = parse_time(str(r.available_from))
            end = parse_time(str(r.available_to))
            if start is None or end is None:
                print(f"⚠️ Skipping TEACHER_AVAILABILITY row with invalid time: {r}")
                continue
            teacher_availability.setdefault(teacher, []).append((day, start, end))

    return {
        "start_date": start_date,
        "end_date": end_date,
        "timeslots": timeslots,
        "requirements": reqs,
        "days": days,
        "breaks": breaks,
        "teacher_availability": teacher_availability
    }

def is_teacher_available(teacher_availability, teacher: str, day: str, slot_start, slot_end) -> bool:
    """
    Check if a teacher is available on a given day and timeslot.
    """
    # If no availability specified for this teacher → assume always available
    availability = [a for a in teacher_availability if a["teacher"] == teacher and a["day"] == day]
    if not availability:
        return True  

    for a in availability:
        if slot_start >= a["available_from"] and slot_end <= a["available_to"]:
            return True
    return False


def export_to_template(assignments, engine, start_date, end_date, output_xlsx, template_xlsx, break_start="12:00", break_end="12:30", holidays=None):
    if holidays is None:
        holidays = []  # list of days like ["Sunday"]

    # Build rows grouped by course-section (up to 5 meetings per row)
    day_rank = {d: i for i, d in enumerate(engine.days)}
    from collections import defaultdict
    buckets = defaultdict(list)

    # Normal assignments
    for (course, section, _), (day, slot_id, room) in assignments.items():
        ts = engine.ts_by_id[slot_id]
        req = engine.req_index[(course, section)]
        buckets[(course, section)].append({
            "course": course,
            "section": section,
            "teacher": req.teacher,
            "curriculum": req.curriculum,
            "semester": req.semester,
            "day": day,
            "time_from": ts.start.strftime("%H:%M:%S"),
            "time_to": ts.end.strftime("%H:%M:%S"),
            "room": room or "",
            "start": ts.start.strftime("%H:%M")
        })

    # Inject break slots (for all courses/sections, common for each day)
    sections = {(req.curriculum, req.semester, req.section_id) for req in engine.requirements}

    for (curr, sem, sec) in sections:
        for day in engine.days:
            if day in holidays:
                continue
            # bucket key without course
            buckets[(None, sec)].append({
                "course": "",   # no course
                "section": sec,
                "teacher": "",  # no teacher
                "curriculum": curr,
                "semester": sem,
                "day": day,
                "time_from": f"{break_start}:00",
                "time_to": f"{break_end}:00",
                "room": "BREAK",
                "start": break_start
            })


    # Build final rows
    rows = []
    for key, items in buckets.items():
        items.sort(key=lambda x: (day_rank.get(x["day"], 999), x["start"]))
        for off in range(0, len(items), 5):
            chunk = items[off:off + 5]
            base = chunk[0]
            row = {
                "STARTDATE": start_date, "ENDDATE": end_date,
                "CURRICULUM": base["curriculum"], "COURSE": base["course"],
                "SEMESTER": base["semester"], "SECTION": base["section"], "TEACHER": base["teacher"],
            }
            for i in range(5):
                dcol, tfcol, ttcol, rcol, lcol = (
                    f"DAY{i+1}", f"DAY{i+1}_TIME_FROM", f"DAY{i+1}_TIME_TO", f"DAY{i+1}_ROOM", f"DAY{i+1}_LINK"
                )
                if i < len(chunk):
                    e = chunk[i]
                    row[dcol], row[tfcol], row[ttcol], row[rcol], row[lcol] = (
                        e["day"], e["time_from"], e["time_to"], e["room"], ""
                    )
                else:
                    row[dcol] = row[tfcol] = row[ttcol] = row[rcol] = row[lcol] = ""
            rows.append(row)

    out_df = pd.DataFrame(rows, columns=[
        "STARTDATE", "ENDDATE", "CURRICULUM", "COURSE", "SEMESTER", "SECTION", "TEACHER",
        "DAY1", "DAY1_TIME_FROM", "DAY1_TIME_TO", "DAY1_ROOM", "DAY1_LINK",
        "DAY2", "DAY2_TIME_FROM", "DAY2_TIME_TO", "DAY2_ROOM", "DAY2_LINK",
        "DAY3", "DAY3_TIME_FROM", "DAY3_TIME_TO", "DAY3_ROOM", "DAY3_LINK",
        "DAY4", "DAY4_TIME_FROM", "DAY4_TIME_TO", "DAY4_ROOM", "DAY4_LINK",
        "DAY5", "DAY5_TIME_FROM", "DAY5_TIME_TO", "DAY5_ROOM", "DAY5_LINK"
    ])

    wb = load_workbook(template_xlsx)
    ws = wb["TimeTable"]
    headers = [c.value for c in ws[1]]

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for _, r in out_df.iterrows():
        ws.append([r.get(h, "") for h in headers])

    wb.save(output_xlsx)
    return output_xlsx

# ------------------- Main Execution -------------------

# Upload your v2 INPUT workbook (must have sheets: WINDOW, TIMESLOTS, REQUIREMENTS, DAYS)
# from google.colab import files
import pandas as pd
import os

REQUIRED_SHEETS = ["WINDOW", "TIMESLOTS", "REQUIREMENTS", "DAYS"]

# input_xlsx = "InputData_v2.xlsx"
input_xlsx = "InputData_Original.xlsx"

print("Using:", input_xlsx)

# Read breaks sheet
breaks_df = pd.read_excel(input_xlsx, sheet_name="BREAKS")

# Function to fetch break time
def get_break_time(curriculum, semester, section_id, day):
    match = breaks_df[
        (breaks_df["curriculum"] == curriculum) &
        (breaks_df["semester"] == semester) &
        (breaks_df["section_id"] == section_id) &
        (breaks_df["day"] == day)
    ]
    if not match.empty:
        return match.iloc[0]["break_from"], match.iloc[0]["break_to"]
    return None, None

# Quick validation
# ✅ Quick validation
if not os.path.exists(input_xlsx):
    raise FileNotFoundError(f"❌ File not found: {input_xlsx}")
    
xls = pd.ExcelFile(input_xlsx)
missing = [s for s in REQUIRED_SHEETS if s not in xls.sheet_names]

if missing:
    raise ValueError(
        f"❌ Input file is missing required sheets: {missing}. "
        f"Found sheets: {xls.sheet_names}"
    )
else:
    print("✅ All required sheets are present!")


# (Optional) quick peek so you can confirm you uploaded the right file
# display(pd.read_excel(input_xlsx, "WINDOW").head())
# display(pd.read_excel(input_xlsx, "TIMESLOTS").head())
# display(pd.read_excel(input_xlsx, "REQUIREMENTS").head())
# display(pd.read_excel(input_xlsx, "DAYS").head())


# Now run the pipeline with this uploaded input
data = read_input_v2(input_xlsx)

engine = TimetableCSPv2(
    data["timeslots"], 
    data["requirements"], 
    data["days"],
    data['teacher_availability']  # <-- pass teacher availability here
)

print("Generating timetable... This may take a few seconds.")


assignments = engine.solve(seed=123)

# Excel template path
template_xlsx = "TimeTableImport_SIS.xlsx"


if not os.path.exists(template_xlsx):
    raise FileNotFoundError(f"❌ Template file not found: {template_xlsx}")

# 🔹 Output file path
out_path = "TimeTableImport_v2.xlsx"

# Export timetable
export_to_template(
    assignments,
    engine,
    data["start_date"],
    data["end_date"],
    out_path,
    template_xlsx
)

print(f"✅ Done! Timetable exported successfully to: {out_path}")




